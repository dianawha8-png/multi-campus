"""
Service d'import Excel.

- Lit un ou plusieurs fichiers .xlsx
- Parcourt automatiquement toutes les feuilles
- Détecte le tableau structuré de chaque feuille (ou, à défaut, la première
  ligne non vide comme en-tête)
- Extrait CAMPUS / SEMESTRE depuis le nom du fichier (convention CAMPUS-SEMESTRE-NOM.xlsx)
- Calcule la cle_import et fait un upsert (INSERT / UPDATE) dans Supabase
- Journalise chaque fichier et chaque erreur

Convention de nom de fichier attendue :
    ABIDJAN-S1-KOUADIO.xlsx  ->  campus="ABIDJAN", semestre="S1", enseignant="KOUADIO"
    YOPOUGON-S4-YAO.xlsx     ->  campus="YOPOUGON", semestre="S4", enseignant="YAO"
Si le nom ne suit pas cette convention, campus/semestre/enseignant sont laissés
vides et doivent être déduits de la colonne CAMPUS ou saisis manuellement.
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import pandas as pd

from services.validation import (
    RapportValidation,
    verifier_colonnes,
    ligne_est_vide,
    matricule_absent,
)

NOM_FICHIER_RE = re.compile(r"^([A-ZÀ-Ÿ]+)-([A-Z0-9]+)-(.+?)\.xlsx?$", re.IGNORECASE)


def parser_nom_fichier(nom_fichier: str) -> dict:
    """Extrait campus / semestre / enseignant depuis le nom du fichier."""
    m = NOM_FICHIER_RE.match(nom_fichier.strip())
    if not m:
        return {"campus": None, "semestre": None, "enseignant": None}
    campus, semestre, enseignant = m.groups()
    return {
        "campus": campus.strip().upper(),
        "semestre": semestre.strip().upper(),
        "enseignant": enseignant.strip().upper(),
    }


def calculer_cle_import(matricule: str, nom: str, prenoms: str, classe: str) -> str:
    matricule = (matricule or "").strip()
    nom = (nom or "").strip().upper()
    prenoms = (prenoms or "").strip().upper()
    classe = (classe or "").strip().upper()
    prefixe = matricule if matricule else "*"
    return f"{prefixe}|{nom}|{prenoms}|{classe}"


def _lire_feuille_en_dataframe(ws) -> pd.DataFrame | None:
    """Lit une feuille openpyxl et retourne un DataFrame à partir de la première
    ligne non vide considérée comme en-tête. Retourne None si la feuille est vide."""
    lignes = list(ws.iter_rows(values_only=True))
    # Supprime les lignes totalement vides en tête
    while lignes and all(c is None for c in lignes[0]):
        lignes.pop(0)
    if not lignes:
        return None

    entete = [str(c).strip().upper() if c is not None else "" for c in lignes[0]]
    donnees = lignes[1:]
    if not donnees:
        return pd.DataFrame(columns=entete)

    df = pd.DataFrame(donnees, columns=entete)
    # Supprime les colonnes sans nom (artefacts de mise en forme Excel)
    df = df.loc[:, [c for c in df.columns if c]]
    return df


def extraire_lignes_fichier(
    nom_fichier: str, contenu: bytes, rapport: RapportValidation
) -> tuple[list[dict], int]:
    """Extrait toutes les lignes valides de toutes les feuilles d'un fichier Excel.

    Retourne (lignes_extraites, nb_feuilles_lues).
    Chaque ligne extraite est un dict prêt pour l'upsert Supabase.
    """
    infos_fichier = parser_nom_fichier(nom_fichier)
    lignes_sortie: list[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(contenu), data_only=True, read_only=True)
    except Exception as exc:
        rapport.ajouter(nom_fichier, "-", 0, "fichier_vide", f"Fichier illisible : {exc}")
        return [], 0

    if not wb.sheetnames:
        rapport.ajouter(nom_fichier, "-", 0, "fichier_vide", "Aucune feuille dans le classeur")
        return [], 0

    nb_feuilles_lues = 0
    for nom_feuille in wb.sheetnames:
        ws = wb[nom_feuille]
        df = _lire_feuille_en_dataframe(ws)

        if df is None or df.empty:
            rapport.ajouter(nom_fichier, nom_feuille, 0, "feuille_vide", "Feuille vide")
            continue

        manquantes = verifier_colonnes(list(df.columns))
        if manquantes:
            rapport.ajouter(
                nom_fichier, nom_feuille, 0, "colonnes_manquantes",
                f"Colonnes manquantes : {', '.join(manquantes)}",
            )
            continue

        nb_feuilles_lues += 1

        for idx, row in df.iterrows():
            ligne_excel = idx + 2  # +1 pour l'en-tête, +1 pour index 0-based
            ligne = row.to_dict()

            if ligne_est_vide(ligne):
                rapport.ajouter(nom_fichier, nom_feuille, ligne_excel, "ligne_vide", "Ligne vide ignorée")
                continue

            if matricule_absent(ligne):
                rapport.ajouter(
                    nom_fichier, nom_feuille, ligne_excel, "matricule_absent",
                    "Matricule absent — clé basée sur nom/prénoms/classe",
                )

            classe = str(ligne.get("CLASSE") or nom_feuille).strip()
            campus = str(ligne.get("CAMPUS") or infos_fichier["campus"] or "").strip().upper()
            matricule = str(ligne.get("MATRICULE") or "").strip()
            nom = str(ligne.get("NOM") or "").strip().upper()
            prenoms = str(ligne.get("PRENOMS") or "").strip().upper()
            ue = str(ligne.get("UE") or "").strip().upper()

            note1 = _to_float(ligne.get("NOTE1"))
            note2 = _to_float(ligne.get("NOTE2"))
            moyenne = _to_float(ligne.get("MOYENNE"))
            if moyenne is None and (note1 is not None or note2 is not None):
                valeurs = [v for v in (note1, note2) if v is not None]
                moyenne = round(sum(valeurs) / len(valeurs), 4) if valeurs else None

            cle_import = calculer_cle_import(matricule, nom, prenoms, classe)

            lignes_sortie.append({
                "matricule": matricule or None,
                "nom": nom,
                "prenoms": prenoms,
                "classe_texte": classe,
                "campus_texte": campus,
                "semestre_texte": infos_fichier["semestre"],
                "ue": ue,
                "enseignant": infos_fichier["enseignant"],
                "note1": note1,
                "note2": note2,
                "moyenne": moyenne,
                "nom_fichier_source": nom_fichier,
                "cle_import": cle_import,
                "derniere_modification": datetime.now(timezone.utc).isoformat(),
            })

    wb.close()
    return lignes_sortie, nb_feuilles_lues


def _to_float(valeur):
    if valeur is None or valeur == "":
        return None
    try:
        return round(float(valeur), 4)
    except (TypeError, ValueError):
        return None


def importer_lignes_supabase(client, lignes: list[dict]) -> dict:
    """Upsert des lignes dans notes_centrales. Clé de conflit : (cle_import, ue).

    Retourne un résumé {inserees, mises_a_jour, erreurs_supabase}.
    """
    if not lignes:
        return {"inserees": 0, "mises_a_jour": 0, "erreurs_supabase": 0}

    # Récupère les clés déjà existantes pour distinguer insert / update
    cles = list({(l["cle_import"], l["ue"]) for l in lignes})
    existantes = set()
    lot = 200
    for i in range(0, len(cles), lot):
        sous_lot = cles[i:i + lot]
        cle_import_vals = list({c for c, _ in sous_lot})
        res = (
            client.table("notes_centrales")
            .select("cle_import, ue")
            .in_("cle_import", cle_import_vals)
            .execute()
        )
        for row in res.data:
            existantes.add((row["cle_import"], row["ue"]))

    inserees, mises_a_jour, erreurs = 0, 0, 0
    lot_taille = 500
    for i in range(0, len(lignes), lot_taille):
        sous_lignes = lignes[i:i + lot_taille]
        try:
            client.table("notes_centrales").upsert(
                sous_lignes, on_conflict="cle_import,ue"
            ).execute()
            for l in sous_lignes:
                if (l["cle_import"], l["ue"]) in existantes:
                    mises_a_jour += 1
                else:
                    inserees += 1
        except Exception:
            erreurs += len(sous_lignes)

    return {"inserees": inserees, "mises_a_jour": mises_a_jour, "erreurs_supabase": erreurs}


def journaliser_import(client, nom_fichier: str, infos_fichier: dict, resume: dict, nb_feuilles: int, rapport: RapportValidation) -> str | None:
    """Enregistre le fichier importé dans fichiers_import et les erreurs dans journal_erreurs."""
    try:
        res = client.table("fichiers_import").insert({
            "nom_fichier": nom_fichier,
            "campus_detecte": infos_fichier.get("campus"),
            "semestre_detecte": infos_fichier.get("semestre"),
            "nb_feuilles": nb_feuilles,
            "nb_lignes_lues": resume.get("inserees", 0) + resume.get("mises_a_jour", 0),
            "nb_lignes_inserees": resume.get("inserees", 0),
            "nb_lignes_maj": resume.get("mises_a_jour", 0),
            "nb_erreurs": rapport.nb_erreurs,
        }).execute()
        fichier_id = res.data[0]["id"] if res.data else None
    except Exception:
        return None

    if fichier_id and rapport.erreurs:
        lignes_erreurs = [{
            "fichier_import_id": fichier_id,
            "nom_fichier": e.nom_fichier,
            "feuille": e.feuille,
            "ligne_excel": e.ligne_excel,
            "type_erreur": e.type_erreur,
            "detail": e.detail,
        } for e in rapport.erreurs]
        try:
            client.table("journal_erreurs").insert(lignes_erreurs).execute()
        except Exception:
            pass

    return fichier_id
