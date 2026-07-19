"""
Export des données affichées vers Excel, CSV ou PDF.
"""
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def vers_excel(df: pd.DataFrame, nom_feuille: str = "Export") -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nom_feuille[:31])
        ws = writer.sheets[nom_feuille[:31]]
        en_tete_police = Font(bold=True, color="FFFFFF")
        en_tete_fond = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = en_tete_police
            cell.fill = en_tete_fond
        for colonne in ws.columns:
            largeur_max = max((len(str(c.value)) for c in colonne if c.value is not None), default=10)
            ws.column_dimensions[colonne[0].column_letter].width = min(largeur_max + 2, 40)
    return buffer.getvalue()


def vers_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def vers_pdf(df: pd.DataFrame, titre: str = "Export") -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(titre, styles["Title"]), Spacer(1, 12)]

    donnees = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(donnees, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()
